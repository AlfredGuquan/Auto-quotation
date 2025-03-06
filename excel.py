import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
import pprint

def extract_cell_styles(template_file, range_str):
    """
    从Excel模板文件中提取指定范围单元格的样式
    
    参数:
    template_file -- 模板文件路径
    range_str -- 单元格范围，如'A1:F22'
    
    返回:
    包含所有单元格样式的字典
    """
    wb = openpyxl.load_workbook(template_file)
    ws = wb.active
    
    # 解析范围
    start_cell, end_cell = range_str.split(':')
    start_col = ord(start_cell[0]) - ord('A') + 1
    start_row = int(start_cell[1:])
    end_col = ord(end_cell[0]) - ord('A') + 1
    end_row = int(end_cell[1:])
    
    styles = {}
    
    for row in range(start_row, end_row + 1):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row, column=col)
            cell_addr = f"{chr(64 + col)}{row}"
            
            # 提取单元格样式
            cell_style = {
                'value': cell.value,
                'font': {
                    'name': cell.font.name,
                    'size': cell.font.size,
                    'bold': cell.font.bold,
                    'italic': cell.font.italic,
                    'color': cell.font.color.rgb if cell.font.color else None
                },
                'alignment': {
                    'horizontal': cell.alignment.horizontal,
                    'vertical': cell.alignment.vertical,
                    'wrap_text': cell.alignment.wrap_text
                },
                'fill': {
                    'fill_type': cell.fill.fill_type,
                    'start_color': cell.fill.start_color.rgb if cell.fill.start_color else None,
                    'end_color': cell.fill.end_color.rgb if cell.fill.end_color else None
                },
                'border': {
                    'left': {
                        'style': cell.border.left.style if cell.border.left else None,
                        'color': cell.border.left.color.rgb if cell.border.left and cell.border.left.color else None
                    },
                    'right': {
                        'style': cell.border.right.style if cell.border.right else None,
                        'color': cell.border.right.color.rgb if cell.border.right and cell.border.right.color else None
                    },
                    'top': {
                        'style': cell.border.top.style if cell.border.top else None,
                        'color': cell.border.top.color.rgb if cell.border.top and cell.border.top.color else None
                    },
                    'bottom': {
                        'style': cell.border.bottom.style if cell.border.bottom else None,
                        'color': cell.border.bottom.color.rgb if cell.border.bottom and cell.border.bottom.color else None
                    }
                }
            }
            
            # 检查合并单元格
            for merged_cell in ws.merged_cells.ranges:
                if cell.coordinate in merged_cell:
                    cell_style['merged'] = str(merged_cell)
                    break
            
            styles[cell_addr] = cell_style
    
    return styles

def generate_style_code(styles):
    """
    生成应用样式的Python代码
    
    参数:
    styles -- 包含单元格样式的字典
    
    返回:
    Python代码字符串
    """
    code = """
def apply_template_styles(ws):
    # 定义常用样式
    thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
    
    # 应用合并单元格
"""
    
    # 处理合并单元格
    merged_cells = {}
    for cell_addr, style in styles.items():
        if 'merged' in style:
            merged_cells[style['merged']] = True
    
    for merged_range in merged_cells:
        code += f"    ws.merge_cells('{merged_range}')\n"
    
    code += "\n    # 应用单元格样式\n"
    
    # 处理单元格样式
    for cell_addr, style in styles.items():
        code += f"    # 设置 {cell_addr} 单元格样式\n"
        
        # 设置值
        if style['value'] is not None:
            if isinstance(style['value'], str):
                # 处理多行文本
                value = style['value'].replace('\n', '\\n').replace("'", "\\'")
                code += f"    ws['{cell_addr}'].value = '{value}'\n"
            else:
                # 处理公式和数值
                if isinstance(style['value'], str) and style['value'].startswith('='):
                    # 这是公式
                    code += f"    ws['{cell_addr}'].value = '{style['value']}'\n"
                else:
                    code += f"    ws['{cell_addr}'].value = {style['value']}\n"
        
        # 设置字体
        font_props = []
        if style['font']['name']:
            font_props.append(f"name='{style['font']['name']}'")
        if style['font']['size']:
            font_props.append(f"size={style['font']['size']}")
        if style['font']['bold']:
            font_props.append(f"bold={style['font']['bold']}")
        if style['font']['italic']:
            font_props.append(f"italic={style['font']['italic']}")
        if style['font']['color'] and isinstance(style['font']['color'], str):
            font_props.append(f"color='{style['font']['color']}'")
        
        if font_props:
            code += f"    ws['{cell_addr}'].font = Font({', '.join(font_props)})\n"
        
        # 设置对齐
        align_props = []
        if style['alignment']['horizontal']:
            align_props.append(f"horizontal='{style['alignment']['horizontal']}'")
        if style['alignment']['vertical']:
            align_props.append(f"vertical='{style['alignment']['vertical']}'")
        if style['alignment']['wrap_text']:
            align_props.append(f"wrap_text={style['alignment']['wrap_text']}")
        
        if align_props:
            code += f"    ws['{cell_addr}'].alignment = Alignment({', '.join(align_props)})\n"
        
        # 设置填充
        if style['fill']['fill_type'] and style['fill']['fill_type'].lower() != 'none' and style['fill']['start_color']:
            start_color = f"'{style['fill']['start_color']}'" if isinstance(style['fill']['start_color'], str) else 'None'
            end_color = f"'{style['fill']['end_color']}'" if style['fill']['end_color'] and isinstance(style['fill']['end_color'], str) else start_color
            code += f"    ws['{cell_addr}'].fill = PatternFill(fill_type='{style['fill']['fill_type']}', start_color={start_color}, end_color={end_color})\n"
        
        # 设置边框
        border_sides = []
        for side in ['left', 'right', 'top', 'bottom']:
            if style['border'][side]['style']:
                color_str = ""
                if style['border'][side]['color'] and isinstance(style['border'][side]['color'], str):
                    color_str = f", color='{style['border'][side]['color']}'"
                border_sides.append(f"{side}=Side(style='{style['border'][side]['style']}'{color_str})")
        
        if border_sides:
            code += f"    ws['{cell_addr}'].border = Border({', '.join(border_sides)})\n"
        
        code += "\n"
    
    return code

def main():
    template_file = "测试.xlsx"  # 替换为您的模板文件路径
    cell_range = "A1:F22"  # 替换为您需要提取的单元格范围
    
    print(f"从 {template_file} 提取 {cell_range} 范围的单元格样式...")
    styles = extract_cell_styles(template_file, cell_range)
    
    code = generate_style_code(styles)
    
    # 保存到文件
    output_file = "template_styles.py"
    with open(output_file, 'w', encoding='utf-8') as f:
        f.write(code)
    
    print(f"样式代码已生成并保存到 {output_file}")
    print("请将生成的代码复制到 generate.py 中使用")

if __name__ == "__main__":
    main()

