import re
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.drawing.image import Image
from datetime import datetime
import os
import math

class ERobPriceCalculator:
    def __init__(self):
        # 基础价格表 - 根据型号和数量范围
        self.price_table = {
            'eRob70F': {
                'retail': 1032.00,
                '5-10': 938.00,
                '10-99': 893.00,
                '100-499': 851.00,
                '500-999': 810.00,
                '1000-1999': 772.00,
                '2000+': 670.00
            },
            'eRob70H': {
                'retail': 1011.00,
                '5-10': 919.00,
                '10-99': 876.00,
                '100-499': 834.00,
                '500-999': 794.00,
                '1000-1999': 756.00,
                '2000+': 657.00
            },
            'eRob80H': {
                'retail': 1072.00,
                '5-10': 974.00,
                '10-99': 928.00,
                '100-499': 884.00,
                '500-999': 842.00,
                '1000-1999': 802.00,
                '2000+': 696.00
            },
            'eRob90H': {
                'retail': 1215.00,
                '5-10': 1104.00,
                '10-99': 1052.00,
                '100-499': 1002.00,
                '500-999': 954.00,
                '1000-1999': 908.00,
                '2000+': 789.00
            },
            'eRob110H': {
                'retail': 1272.00,
                '5-10': 1156.00,
                '10-99': 1101.00,
                '100-499': 1049.00,
                '500-999': 999.00,
                '1000-1999': 951.00,
                '2000+': 826.00
            },
            'eRob142H': {
                'retail': 1338.00,
                '5-10': 1216.00,
                '10-99': 1158.00,
                '100-499': 1103.00,
                '500-999': 1050.00,
                '1000-1999': 1000.00,
                '2000+': 868.00
            },
            'eRob170H': {
                'retail': 2008.00,
                '5-10': 1826.00,
                '10-99': 1739.00,
                '100-499': 1656.00,
                '500-999': 1577.00,
                '1000-1999': 1502.00,
                '2000+': 1304.00
            }
        }
        
        # 选项价格表
        self.option_prices = {
            'without_brake': -15,
            'multiturn': 23,
            'ethercat': 23,
            'high_precision': 46,
            't_type': {
                70: 77,
                80: 77,
                90: 77,
                110: 138,
                142: 200,
                170: 200
            },
            'version': {
                'V6': {
                    80: 78,
                    90: 180,
                    110: 92
                }
            },
            'leaderdrive_gear': {
                70: 83,
                80: 111,
                90: 129,
                110: 138,
                142: 258,
                170: 277
            }
        }
        
        # 默认减速比映射
        self.default_ratio = {
            70: 100,
            80: 120,
            90: 120,
            110: 120,
            142: 120,
            170: 120
        }
        
        # 产品重量（kg）
        self.weights = {
            70: 0.85,
            80: 1.03,
            90: 1.62,
            110: 2.68,
            142: 4.50,
            170: 7.20,
            'eRob Universal Accessories Kit': 0.2,
            'eLine - RJ45 ECAT -30': 0.03,
            'eRob to PC Connector': 0.2
        }
        
        # 配件价格
        self.accessories = {
            'eRob Universal Accessories Kit': 35,
            'eLine - RJ45 ECAT -30': 11,
            'eRob to PC Connector': 66
        }
        
        # 产品图片路径
        self.images = {
            # eRob系列产品图片
            'eRob70F': 'images/70F.png',
            'eRob70H': 'images/70I.png',
            'eRob80H': 'images/80I.png',
            'eRob90H': 'images/90I.png',
            'eRob110H': 'images/110I.png',
            'eRob142H': 'images/142I.png',
            'eRob170H': 'images/170I.png',
            # 配件图片
            'eRob Universal Accessories Kit': 'images/Kit.png',
            'eLine - RJ45 ECAT -30': 'images/RJ45.png',
            'eRob to PC Connector': 'images/PC.png'
        }
    
    def normalize_model_code(self, model_code):
        """标准化型号编码，补全缺失部分"""
        # 处理特殊版本标记
        version = ""
        if '[' in model_code:
            base_code, version = model_code.split('[')
            version = version.strip(']')
            model_code = base_code.strip()
        
        # 如果不是以eRob开头，添加eRob前缀
        if not model_code.startswith('eRob'):
            # 检查是否以数字开头（如80H120I-BHM-18CN）
            if model_code[0].isdigit():
                model_code = 'eRob' + model_code
        
        # 解析型号主体部分
        parts = model_code.split('-')
        base_info = parts[0]
        
        # 提取直径和齿轮类型
        if 'eRob' in base_info:
            diameter_start = base_info.find('eRob') + 4
            diameter_end = diameter_start
            while diameter_end < len(base_info) and base_info[diameter_end].isdigit():
                diameter_end += 1
            
            if diameter_end > diameter_start:
                diameter = int(base_info[diameter_start:diameter_end])
                gear_type = base_info[diameter_end] if diameter_end < len(base_info) else 'H'  # 默认H型
            else:
                # 无法提取直径，使用默认值
                diameter = 80
                gear_type = 'H'
        else:
            # 无法识别eRob，使用默认值
            diameter = 80
            gear_type = 'H'
        
        # 检查是否包含减速比
        ratio_match = re.search(r'(\d+)[IT]', base_info)
        if not ratio_match:
            # 添加默认减速比
            default_ratio = self.default_ratio.get(diameter, 120)
            if 'I' in base_info or 'T' in base_info:
                # 已有形状标识，在其前添加减速比
                form_type_pos = base_info.find('I') if 'I' in base_info else base_info.find('T')
                base_info = base_info[:form_type_pos] + str(default_ratio) + base_info[form_type_pos:]
            else:
                # 无形状标识，添加减速比和默认I型
                base_info = base_info + str(default_ratio) + 'I'
        
        # 确保有形状标识（I或T）
        if 'I' not in base_info and 'T' not in base_info:
            base_info = base_info + 'I'  # 默认I型
        
        # 重建完整型号
        normalized_code = base_info
        if len(parts) > 1:
            normalized_code += '-' + parts[1]
        else:
            normalized_code += '-BHM'  # 默认配置
        
        if len(parts) > 2:
            normalized_code += '-' + parts[2]
        else:
            normalized_code += '-18CN'  # 默认接口
        
        # 添加版本标记（如果有）
        if version:
            normalized_code += f"[{version}]"
        
        return normalized_code
    
    def parse_model_code(self, model_code):
        """解析型号编码，提取各部分信息"""
        # 标准化型号
        normalized_model = self.normalize_model_code(model_code)
        
        # 处理特殊版本标记
        version = ""
        if '[' in normalized_model:
            base_model, version_part = normalized_model.split('[')
            version = version_part.strip(']')
            normalized_model = base_model.strip()
        
        # 解析型号主体部分
        parts = normalized_model.split('-')
        base_info = parts[0]
        
        # 提取直径和齿轮类型
        diameter_start = base_info.find('eRob') + 4
        diameter_end = diameter_start
        while diameter_end < len(base_info) and base_info[diameter_end].isdigit():
            diameter_end += 1
        
        diameter = int(base_info[diameter_start:diameter_end])
        gear_type = base_info[diameter_end]
        
        # 提取减速比
        ratio_match = re.search(r'(\d+)[IT]', base_info)
        ratio = int(ratio_match.group(1)) if ratio_match else self.default_ratio.get(diameter, 120)
        
        # 提取形状类型
        form_type = 'I'  # 默认I型
        if 'T' in base_info[base_info.find(str(ratio)) + len(str(ratio)):]:
            form_type = 'T'
        
        # 解析配置部分
        config = parts[1] if len(parts) > 1 else ''
        
        # 解析通信和传感器部分
        interface = parts[2] if len(parts) > 2 else ''
        
        return {
            'full_model': model_code + (f"[{version}]" if version else ""),
            'base_model': f"eRob{diameter}{gear_type}",
            'diameter': diameter,
            'gear_type': gear_type,
            'ratio': ratio,
            'form_type': form_type,
            'config': config,
            'interface': interface,
            'version': version,
            'has_ethercat': 'E' in interface
        }
    
    def get_price_range(self, quantity):
        """根据数量确定价格范围"""
        if quantity < 5:
            return 'retail'
        elif quantity < 10:
            return '5-10'
        elif quantity < 100:
            return '10-99'
        elif quantity < 500:
            return '100-499'
        elif quantity < 1000:
            return '500-999'
        elif quantity < 2000:
            return '1000-1999'
        else:
            return '2000+'
    
    def calculate_price(self, model_code, quantity):
        """计算指定型号和数量的价格"""
        # 解析型号
        model_info = self.parse_model_code(model_code)
        
        # 获取基础价格
        price_range = self.get_price_range(quantity)
        base_price = self.price_table.get(model_info['base_model'], {}).get(price_range, 0)
        
        # 计算选项价格
        options_price = 0
        
        # 制动器选项
        if 'F' in model_info['config']:  # 无制动器
            options_price += self.option_prices['without_brake']
        
        # 多圈编码器
        if 'M' in model_info['config']:
            options_price += self.option_prices['multiturn']
        
        # 高精度编码器
        if 'H' in model_info['config']:
            options_price += self.option_prices['high_precision']
        
        # 通信协议
        if 'E' in model_info['interface']:  # EtherCAT
            options_price += self.option_prices['ethercat']
        
        # T型结构
        if model_info['form_type'] == 'T':
            t_price = self.option_prices['t_type'].get(model_info['diameter'], 77)
            options_price += t_price
        
        # 版本加价
        if model_info['version']:
            if model_info['version'].upper() == 'V6' and model_info['diameter'] in [80, 90, 110]:
                version_price = self.option_prices['version']['V6'].get(model_info['diameter'], 0)
                options_price += version_price
        
        # 计算总价
        unit_price = base_price + options_price
        total_price = unit_price * quantity
        
        # 获取重量
        weight = self.weights.get(model_info['diameter'], 1.0)
        
        return {
            'model': model_info['full_model'],
            'normalized_model': self.normalize_model_code(model_code),
            'base_price': base_price,
            'options_price': options_price,
            'unit_price': unit_price,
            'quantity': quantity,
            'total_price': total_price,
            'weight': weight,
            'details': {
                'base_model': model_info['base_model'],
                'diameter': model_info['diameter'],
                'price_range': price_range,
                'without_brake': 'F' in model_info['config'],
                'multiturn': 'M' in model_info['config'],
                'high_precision': 'H' in model_info['config'],
                'ethercat': 'E' in model_info['interface'],
                'form_type': model_info['form_type'],
                'version': model_info['version']
            }
        }
    
    def calculate_batch(self, model_codes, quantities, customer_info=None):
        """批量计算多个型号的价格并生成完整报价单"""
        if customer_info is None:
            customer_info = {
                'name': 'Customer Name',
                'company': 'Company Name',
                'address': 'Street Address',
                'city': 'City',
                'zip': 'ZIP Code',
                'country': 'Country',
                'phone': 'Phone Number'
            }
        
        results = []
        total_erob_quantity = 0
        has_ethercat = False
        
        for i, model_code in enumerate(model_codes):
            quantity = quantities[i] if i < len(quantities) else 1
            result = self.calculate_price(model_code, quantity)
            results.append(result)
            total_erob_quantity += quantity
            
            # 检查是否有EtherCAT
            model_info = self.parse_model_code(model_code)
            if model_info['has_ethercat']:
                has_ethercat = True
        
        # 添加默认配件
        # 1. eRob Universal Accessories Kit
        accessories_kit = {
            'model': 'eRob Universal Accessories Kit',
            'normalized_model': 'eRob Universal Accessories Kit',
            'unit_price': self.accessories['eRob Universal Accessories Kit'],
            'quantity': total_erob_quantity,
            'total_price': self.accessories['eRob Universal Accessories Kit'] * total_erob_quantity,
            'weight': self.weights['eRob Universal Accessories Kit'],
            'is_accessory': True
        }
        results.append(accessories_kit)
        
        # 2. eLine - RJ45 ECAT -30 (如果有EtherCAT)
        if has_ethercat:
            eline = {
                'model': 'eLine - RJ45 ECAT -30',
                'normalized_model': 'eLine - RJ45 ECAT -30',
                'unit_price': self.accessories['eLine - RJ45 ECAT -30'],
                'quantity': total_erob_quantity,
                'total_price': self.accessories['eLine - RJ45 ECAT -30'] * total_erob_quantity,
                'weight': self.weights['eLine - RJ45 ECAT -30'],
                'is_accessory': True
            }
            results.append(eline)
        
        # 3. eRob to PC Connector
        connector = {
            'model': 'eRob to PC Connector',
            'normalized_model': 'eRob to PC Connector',
            'unit_price': self.accessories['eRob to PC Connector'],
            'quantity': 1,
            'total_price': self.accessories['eRob to PC Connector'],
            'weight': self.weights['eRob to PC Connector'],
            'is_accessory': True
        }
        results.append(connector)
        
        # 计算总价和总重量
        subtotal = sum(item['total_price'] for item in results)
        total_weight = sum(item['weight'] * item['quantity'] for item in results)
        
        # 计算运费 (使用公式 ROUND(4830*1.1/6.5,0))
        freight = round(4830 * 1.1 / 6.5, 0)
        
        # 计算总计
        grand_total = subtotal + freight
        
        return {
            'items': results,
            'subtotal': subtotal,
            'freight': freight,
            'grand_total': grand_total,
            'total_weight': total_weight,
            'customer_info': customer_info,
            'has_ethercat': has_ethercat
        }
    
    def export_to_quotation(self, results):
        """导出报价单到Excel文件"""
        # 创建工作簿
        wb = Workbook()
        ws = wb.active
        ws.title = "Quotation"
        
        # 设置列宽
        ws.column_dimensions['A'].width = 15
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15
        
        # 设置合并单元格
        ws.merge_cells('A1:F1')  # 标题行
        ws.merge_cells('A2:B5')  # 客户信息区
        ws.merge_cells('D2:F5')  # 日期和引用信息区
        ws.merge_cells('A6:F6')  # Quotation List 标题栏
        
        # 设置标题和样式
        cell = ws['A1']
        cell.value = "ZeroErr Control Co., Ltd."
        cell.font = Font(name='Microsoft YaHei', size=16, bold=True, color='002060')
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置表头样式
        cell = ws['A6']
        cell.value = "Quotation List"
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # 设置客户信息
        customer_info = results['customer_info']
        to_text = f"To:  {customer_info.get('name', '')}\n"
        to_text += f"Company: {customer_info.get('company', '')}\n"
        to_text += f"Street Address: {customer_info.get('address', '')}\n"
        to_text += f"City, ST  ZIP Code: {customer_info.get('city', '')}, {customer_info.get('zip', '')}, {customer_info.get('country', '')}\n"
        to_text += f"Phone: {customer_info.get('phone', '')}"
        cell = ws['A2']
        cell.value = to_text
        cell.font = Font(name='Microsoft YaHei', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # 设置日期和报价号
        today = datetime.now()
        date_str = today.strftime('%Y.%m.%d')
        quotation_number = today.strftime('%Y%m%d') + '01'
        
        date_text = f"Date：{date_str}\n"
        date_text += f"Quotation #: {quotation_number}\n"
        date_text += f"Street Address：Fuyuan 1st, Fuhai City, ZIP Code：\n"
        date_text += f"Bao'an，Shen Zhen, 518103\n"
        date_text += f"Phone：+86 18922807806"
        cell = ws['D2']
        cell.value = date_text
        cell.font = Font(name='Microsoft YaHei', size=10)
        cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        
        # 设置表头
        headers = ["IMAGES", "MODELS", "QUANTITY (PC)", "WEIGHT (KG/PC)", "UNIT PRICE (USD/PC)", "AMOUNT (USD)"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=7, column=col)
            cell.value = header
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
        
        # 设置行高
        ws.row_dimensions[7].height = 40  # 表头行高
        
        # 填充产品数据
        row = 8
        for item in results['items']:
            # 设置行高以适应图片
            ws.row_dimensions[row].height = 60
            
            # 为所有单元格添加边框
            for col in range(1, 7):
                cell = ws.cell(row=row, column=col)
                cell.border = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )
            
            # 插入产品图片
            model_key = item.get('normalized_model', item.get('model', ''))
            if model_key in self.images and os.path.exists(self.images[model_key]):
                try:
                    img = Image(self.images[model_key])
                    img.width = 60
                    img.height = 60
                    ws.add_image(img, f'A{row}')
                except Exception as e:
                    print(f"插入图片时出错: {e}")
            
            # 填充产品信息
            ws.cell(row=row, column=2).value = item['normalized_model']
            ws.cell(row=row, column=2).alignment = Alignment(vertical='center')
            
            ws.cell(row=row, column=3).value = item['quantity']
            ws.cell(row=row, column=3).alignment = Alignment(horizontal='center', vertical='center')
            
            ws.cell(row=row, column=4).value = item['weight']
            ws.cell(row=row, column=4).alignment = Alignment(horizontal='center', vertical='center')
            
            ws.cell(row=row, column=5).value = item['unit_price']
            ws.cell(row=row, column=5).alignment = Alignment(horizontal='center', vertical='center')
            
            # 设置金额格式
            cell = ws.cell(row=row, column=6)
            cell.value = f"$ {item['total_price']:,.2f}"
            cell.alignment = Alignment(horizontal='right', vertical='center')
            
            row += 1
        
        # 设置备注和合计区域
        remarks_row = row
        
        # 合并备注单元格
        ws.merge_cells(f'A{remarks_row}:D{remarks_row+6}')
        
        remarks_cell = ws.cell(row=remarks_row, column=1)
        remarks_text = "Remarks:\n"
        remarks_text += f"1. Price term: DAP {customer_info.get('country', 'Australia')}\n"
        remarks_text += "2. Payment term: T/T, 100% advance payment.\n"
        remarks_text += "3. Leading time: 12 working days after the payment.\n"
        remarks_text += "4.The price needs to be updated if the exchange rate fluctuate more than 10%."
        
        remarks_cell.value = remarks_text
        remarks_cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
        remarks_cell.font = Font(name='Microsoft YaHei', size=10)
        
        # 设置合计区域
        subtotal_cell = ws.cell(row=remarks_row, column=5)
        subtotal_cell.value = "SUBTOTAL"
        subtotal_cell.font = Font(bold=True)
        subtotal_cell.alignment = Alignment(horizontal='center', vertical='center')
        subtotal_cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        subtotal_amount = ws.cell(row=remarks_row, column=6)
        subtotal_amount.value = f"$ {results['subtotal']:,.2f}"
        subtotal_amount.alignment = Alignment(horizontal='right', vertical='center')
        subtotal_amount.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        freight_cell = ws.cell(row=remarks_row+1, column=5)
        freight_cell.value = "FREIGHT"
        freight_cell.font = Font(bold=True)
        freight_cell.alignment = Alignment(horizontal='center', vertical='center')
        freight_cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        freight_amount = ws.cell(row=remarks_row+1, column=6)
        freight_amount.value = f"$ {results['freight']:,.2f}"
        freight_amount.alignment = Alignment(horizontal='right', vertical='center')
        freight_amount.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        other_cell = ws.cell(row=remarks_row+2, column=5)
        other_cell.value = "OTHER"
        other_cell.font = Font(bold=True)
        other_cell.alignment = Alignment(horizontal='center', vertical='center')
        other_cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        other_amount = ws.cell(row=remarks_row+2, column=6)
        other_amount.value = "$ -"
        other_amount.alignment = Alignment(horizontal='center', vertical='center')
        other_amount.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        total_cell = ws.cell(row=remarks_row+3, column=5)
        total_cell.value = "TOTAL"
        total_cell.font = Font(bold=True)
        total_cell.alignment = Alignment(horizontal='center', vertical='center')
        total_cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        total_amount = ws.cell(row=remarks_row+3, column=6)
        total_amount.value = f"$ {results['grand_total']:,.2f}"
        total_amount.alignment = Alignment(horizontal='right', vertical='center')
        total_amount.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 保存文件
        customer_name = customer_info.get('name', 'Customer').replace(' ', '_')
        filename = f"ZeroErr_Quotation_{customer_name}_{today.strftime('%Y%m%d')}.xlsx"
        wb.save(filename)
        return filename

def main():
    calculator = ERobPriceCalculator()
    
    print("eRob产品报价单生成系统")
    print("请输入客户信息:")
    customer_name = input("客户姓名: ")
    customer_company = input("公司名称: ")
    customer_address = input("街道地址: ")
    customer_city = input("城市: ")
    customer_zip = input("邮编: ")
    customer_country = input("国家: ")
    customer_phone = input("电话: ")
    
    customer_info = {
        'name': customer_name,
        'company': customer_company,
        'address': customer_address,
        'city': customer_city,
        'zip': customer_zip,
        'country': customer_country,
        'phone': customer_phone
    }
    
    print("\n请输入产品型号和数量，多个产品请用逗号分隔")
    print("例如: eRob90H160I-BHM-18ET[V6],80H120I-BHM-18CN")
    print("数量例如: 6,8")
    
    model_input = input("产品型号: ")
    quantity_input = input("对应数量: ")
    
    model_codes = [code.strip() for code in model_input.split(',')]
    quantities = [int(qty.strip()) for qty in quantity_input.split(',')]
    
    # 确保数量与型号数量匹配
    while len(quantities) < len(model_codes):
        quantities.append(1)  # 默认数量为1
    
    results = calculator.calculate_batch(model_codes, quantities, customer_info)
    
    # 打印结果
    print("\n计算结果:")
    print("-" * 80)
    print(f"{'序号':<5}{'产品型号':<30}{'数量':<8}{'单价':<12}{'总价':<12}")
    print("-" * 80)
    
    for i, item in enumerate(results['items'], 1):
        print(f"{i:<5}{item['model']:<30}{item['quantity']:<8}{item['unit_price']:<12.2f}{item['total_price']:<12.2f}")
    
    print("-" * 80)
    print(f"{'小计':<43}{results['subtotal']:<12.2f}")
    print(f"{'运费':<43}{results['freight']:<12.2f}")
    print(f"{'总计':<43}{results['grand_total']:<12.2f}")
    
    # 导出报价单
    excel_file = calculator.export_to_quotation(results)
    print(f"\n报价单已生成: {excel_file}")

if __name__ == "__main__":
    main()